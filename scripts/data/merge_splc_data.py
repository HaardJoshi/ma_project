"""
merge_splc_data.py  --  Parse populated Bloomberg SPLC ALL_DATA Excel files
================================================================================
Reads the 5 Bloomberg-populated Excel workbooks and extracts the full supply
chain data (suppliers + customers with revenue %, cost %, etc.) into a clean CSV.

Keeps only graph-relevant columns:
  - deal_id, acquirer_name, acquirer_ticker
  - role (supplier/customer)
  - entity_ticker, entity_name
  - revenue_pct, cost_pct, relationship_amount
"""

import pandas as pd
from openpyxl import load_workbook
import os
import glob

# ── CONFIG ──────────────────────────────────────────────────────────────────
INPUT_PATTERN = "data/raw/bbg_splc_full_pull_*.xlsx"
DEALS_FILE    = "data/interim/deals_master.csv"
OUTPUT_CSV    = "data/interim/splc_full_data.csv"

COLS_PER_DEAL     = 30
SUPPLIER_OFFSET   = 0    # Supplier BDS starts at col_start + 0
CUSTOMER_OFFSET   = 14   # Customer BDS starts at col_start + 14
DATA_START_ROW    = 5    # Row where BDS data begins

# Column offsets within each 12-column BDS spill block
# SUPPLIERS layout:  Ticker, CostType, Revenue%, Cost%, Period, Year, Source, Date, Count, Amount, CompID, Name
# CUSTOMERS layout:  Ticker, Revenue%, CostType, Cost%, Period, Year, Source, Date, Count, Amount, CompID, Name
# Note: columns 1 and 2 are SWAPPED between suppliers and customers!
SUPPLIER_COL_MAP = {
    "entity": 0, "cost_type": 1, "revenue_pct": 2, "cost_pct": 3,
    "rel_year": 5, "rel_amt": 9, "comp_id": 10, "comp_name": 11
}
CUSTOMER_COL_MAP = {
    "entity": 0, "revenue_pct": 1, "cost_type": 2, "cost_pct": 3,
    "rel_year": 5, "rel_amt": 9, "comp_id": 10, "comp_name": 11
}


def parse_entity_block(ws, col_start, max_row, col_map):
    """Parse a single BDS spill block (supplier or customer) and return records."""
    entities = []
    for row in range(DATA_START_ROW, max_row + 1):
        ticker_val = ws.cell(row=row, column=col_start + col_map["entity"]).value
        if ticker_val is None:
            break
        ticker_str = str(ticker_val).strip()
        if ticker_str in ("", "#N/A N/A", "#N/A", "N/A"):
            break

        cost_type = ws.cell(row=row, column=col_start + col_map["cost_type"]).value
        rev_pct   = ws.cell(row=row, column=col_start + col_map["revenue_pct"]).value
        cost_pct  = ws.cell(row=row, column=col_start + col_map["cost_pct"]).value
        rel_year  = ws.cell(row=row, column=col_start + col_map["rel_year"]).value
        rel_amt   = ws.cell(row=row, column=col_start + col_map["rel_amt"]).value
        comp_id   = ws.cell(row=row, column=col_start + col_map["comp_id"]).value
        comp_name = ws.cell(row=row, column=col_start + col_map["comp_name"]).value

        entities.append({
            "entity_ticker": ticker_str,
            "entity_name": str(comp_name).strip() if comp_name else None,
            "entity_bbg_id": comp_id,
            "revenue_pct": rev_pct,
            "cost_pct": cost_pct,
            "cost_type": str(cost_type).strip() if cost_type else None,
            "relationship_amount": rel_amt,
            "relationship_year": rel_year,
        })
    return entities


def main():
    print("=" * 60)
    print("  PARSING BLOOMBERG SPLC ALL_DATA FILES")
    print("=" * 60)

    # Load deals master
    deals_df = pd.read_csv(DEALS_FILE)
    deal_info = {}
    for _, row in deals_df.iterrows():
        deal_info[row["deal_id"]] = {
            "acquirer_ticker": row["acq_ticker_bbg"],
            "acquirer_name": row.get("acquirer_name", row["acq_ticker_bbg"]),
        }
    print(f"Loaded {len(deal_info):,} deals from {DEALS_FILE}")

    # Find all populated Excel files
    files = sorted(glob.glob(INPUT_PATTERN))
    print(f"Found {len(files)} Excel files to parse.")

    all_records = []
    deals_processed = 0
    deals_with_suppliers = 0
    deals_with_customers = 0

    for filepath in files:
        print(f"\n  Loading {os.path.basename(filepath)}...")
        wb = load_workbook(filepath, data_only=True)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Find deal positions by scanning Row 1
            deal_positions = []
            for col in range(1, ws.max_column + 1):
                val = ws.cell(row=1, column=col).value
                if val and str(val).startswith("Deal "):
                    try:
                        did = int(str(val).replace("Deal ", "").strip())
                        deal_positions.append((did, col))
                    except ValueError:
                        pass

            for did, col_start in deal_positions:
                deals_processed += 1
                info = deal_info.get(did, {})
                acq_ticker = info.get("acquirer_ticker", "UNKNOWN")
                acq_name = info.get("acquirer_name", "UNKNOWN")

                # Parse Suppliers
                sup_entities = parse_entity_block(ws, col_start + SUPPLIER_OFFSET, ws.max_row, SUPPLIER_COL_MAP)
                if sup_entities:
                    deals_with_suppliers += 1
                for ent in sup_entities:
                    ent["deal_id"] = did
                    ent["acquirer_name"] = acq_name
                    ent["acquirer_ticker"] = acq_ticker
                    ent["role"] = "supplier"
                    all_records.append(ent)

                # Parse Customers
                cust_entities = parse_entity_block(ws, col_start + CUSTOMER_OFFSET, ws.max_row, CUSTOMER_COL_MAP)
                if cust_entities:
                    deals_with_customers += 1
                for ent in cust_entities:
                    ent["deal_id"] = did
                    ent["acquirer_name"] = acq_name
                    ent["acquirer_ticker"] = acq_ticker
                    ent["role"] = "customer"
                    all_records.append(ent)

        wb.close()

    # Build DataFrame
    df = pd.DataFrame(all_records)

    # Reorder columns for clarity
    col_order = [
        "deal_id", "acquirer_name", "acquirer_ticker",
        "role", "entity_ticker", "entity_name", "entity_bbg_id",
        "revenue_pct", "cost_pct", "cost_type",
        "relationship_amount", "relationship_year"
    ]
    df = df[col_order]

    # Clean numeric columns
    for col in ["revenue_pct", "cost_pct", "relationship_amount"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    print(f"\n{'='*60}")
    print(f"RESULTS:")
    print(f"  Deals processed:       {deals_processed:,}")
    print(f"  Deals with suppliers:  {deals_with_suppliers:,}")
    print(f"  Deals with customers:  {deals_with_customers:,}")
    print(f"  Total entity records:  {len(df):,}")

    if len(df) > 0:
        suppliers = df[df["role"] == "supplier"]
        customers = df[df["role"] == "customer"]
        print(f"\n  Supplier records:  {len(suppliers):,}")
        print(f"  Customer records:  {len(customers):,}")
        print(f"\n  Revenue % stats (suppliers):")
        print(f"    Mean: {suppliers['revenue_pct'].mean():.2f}%")
        print(f"    Median: {suppliers['revenue_pct'].median():.2f}%")
        print(f"    Non-null: {suppliers['revenue_pct'].notna().sum():,}")
        print(f"\n  Revenue % stats (customers):")
        print(f"    Mean: {customers['revenue_pct'].mean():.2f}%")
        print(f"    Median: {customers['revenue_pct'].median():.2f}%")
        print(f"    Non-null: {customers['revenue_pct'].notna().sum():,}")
        print(f"\nSample data:")
        print(df.head(5).to_string(index=False))

    df.to_csv(OUTPUT_CSV, index=False)
    print(f"\nSaved to {OUTPUT_CSV}")
    print("=" * 60)


if __name__ == "__main__":
    main()
