"""
merge_bbg_data.py  --  Parse Bloomberg-populated Excel and merge into timeseries_long.csv
==========================================================================================
After opening bbg_pull_missing.xlsx in Excel and letting BDH formulas populate,
save the file and run this script. It reads the populated prices, computes log
returns and rel_day, then appends to timeseries_long.csv.
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
import sys

# ── CONFIG (must match generate_bbg_excel.py and pull_car_data.py) ──────────
BBG_EXCEL       = "bbg_pull_missing.xlsx"
DEALS_FILE      = "deals_master.csv"
TS_FILE         = "timeseries_long.csv"
BENCHMARK_LABEL = "SPX Index"

EST_START  = -200
EST_END    =  -20
EVT_START  =   -5
EVT_END    =   +5
MIN_EST_OBS = 120

COLS_PER_DEAL   = 6
DATA_START_ROW  = 5   # 1-indexed (same as generator)


def flag_window(rel_day):
    if EST_START <= rel_day <= EST_END:
        return "EST"
    elif EVT_START <= rel_day <= EVT_END:
        return "EVENT"
    else:
        return "GAP"


def parse_column_pair(ws, col_date, col_price, start_row, max_row):
    """Read a date+price column pair from the worksheet.
    Handles Bloomberg #N/A errors by skipping those rows.
    """
    dates = []
    prices = []
    na_count = 0
    for row in range(start_row, max_row + 1):
        date_val = ws.cell(row=row, column=col_date).value
        price_val = ws.cell(row=row, column=col_price).value
        # Stop at end of data block
        if date_val is None and price_val is None:
            break
        # Skip Bloomberg error values (#N/A, #N/A Invalid Security, etc.)
        if price_val is None or (isinstance(price_val, str) and '#N/A' in str(price_val)):
            na_count += 1
            # If first 5 rows are all #N/A, the security is likely invalid -- stop early
            if na_count >= 5 and len(prices) == 0:
                break
            continue
        try:
            dates.append(pd.Timestamp(date_val))
            prices.append(float(price_val))
        except (ValueError, TypeError):
            continue  # skip unparseable rows, don't break
    return dates, prices


def process_deal(acq_dates, acq_prices, mkt_dates, mkt_prices, ann_date,
                 deal_id, deal_key, bbg_ticker):
    """Process one deal's raw price data into tidy time-series rows.
    Returns a list of DataFrames, or empty list on failure.
    """
    if len(acq_prices) < MIN_EST_OBS + 11 or len(mkt_prices) < MIN_EST_OBS + 11:
        return None, "no_data"

    acq_df = pd.DataFrame({"px_last": acq_prices}, index=pd.DatetimeIndex(acq_dates))
    mkt_df = pd.DataFrame({"px_last": mkt_prices}, index=pd.DatetimeIndex(mkt_dates))

    # Align on common trading days
    common_idx = acq_df.index.intersection(mkt_df.index)
    if len(common_idx) < MIN_EST_OBS + 11:
        return None, "no_data"

    acq_df = acq_df.loc[common_idx].sort_index()
    mkt_df = mkt_df.loc[common_idx].sort_index()

    # Assign rel_day (trading-day based)
    for df in [acq_df, mkt_df]:
        candidates = df.index[df.index >= ann_date]
        if len(candidates) == 0:
            return None, "no_day0"
        day0 = candidates[0]
        day0_pos = df.index.get_loc(day0)
        df["rel_day"] = np.arange(len(df)) - day0_pos

    # Compute log returns
    acq_df["ret_1d"] = np.log(acq_df["px_last"] / acq_df["px_last"].shift(1))
    mkt_df["ret_1d"] = np.log(mkt_df["px_last"] / mkt_df["px_last"].shift(1))
    acq_df = acq_df.dropna(subset=["ret_1d"])
    mkt_df = mkt_df.dropna(subset=["ret_1d"])

    # Flag windows
    acq_df["window_flag"] = acq_df["rel_day"].apply(flag_window)
    mkt_df["window_flag"] = mkt_df["rel_day"].apply(flag_window)

    # Quality check
    n_est = (acq_df["window_flag"] == "EST").sum()
    if n_est < MIN_EST_OBS:
        return None, "few_est"

    # Keep only EST + EVENT rows
    acq_out = acq_df[acq_df["window_flag"].isin(["EST", "EVENT"])].copy()
    mkt_out = mkt_df[mkt_df["window_flag"].isin(["EST", "EVENT"])].copy()

    results = []
    for df_out, role, sec_label in [
        (acq_out, "ACQUIRER", bbg_ticker),
        (mkt_out, "BENCHMARK", BENCHMARK_LABEL),
    ]:
        df_out["deal_id"]       = deal_id
        df_out["deal_key"]      = deal_key
        df_out["security_role"] = role
        df_out["security"]      = sec_label
        df_out["ann_date"]      = ann_date
        df_out.index.name       = "trading_date"
        results.append(df_out.reset_index())

    return results, "ok"


def main():
    print("=" * 70)
    print("  Merge Bloomberg Excel Data")
    print("=" * 70)

    # Load deals master
    deals = pd.read_csv(DEALS_FILE)
    deals["announce_date"] = pd.to_datetime(deals["announce_date"])

    # Load existing time-series
    existing_ts = pd.read_csv(TS_FILE)
    existing_deal_ids = set(existing_ts["deal_id"].unique())

    # Identify missing deals
    missing = deals[~deals["deal_id"].isin(existing_deal_ids)].copy()
    print(f"\nMissing deals to parse: {len(missing)}")

    # Load Bloomberg Excel workbook (data_only=True reads values, not formulas)
    print(f"Loading {BBG_EXCEL} (data_only mode)...")
    wb = load_workbook(BBG_EXCEL, data_only=True)

    ts_rows = []
    success = 0
    skip_no_data = 0
    skip_few_est = 0
    skip_no_day0 = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"  Processing sheet: {sheet_name}...", end="", flush=True)
        sheet_success = 0

        # Scan for deal headers at row 1
        col = 1
        while col <= ws.max_column:
            cell_val = ws.cell(row=1, column=col).value
            if cell_val is None or not str(cell_val).startswith("Deal "):
                col += 1
                continue

            # Extract deal_id
            try:
                deal_id = int(str(cell_val).replace("Deal ", ""))
            except ValueError:
                col += COLS_PER_DEAL
                continue

            # Look up deal info
            deal_row = deals[deals["deal_id"] == deal_id]
            if deal_row.empty:
                col += COLS_PER_DEAL
                continue
            deal = deal_row.iloc[0]
            ann_date = pd.Timestamp(deal["announce_date"])
            bbg_ticker = deal["acq_ticker_bbg"]

            # Parse acquirer price data (cols: col, col+1)
            acq_dates, acq_prices = parse_column_pair(
                ws, col, col + 1, DATA_START_ROW, ws.max_row
            )

            # Parse benchmark price data (cols: col+2, col+3)
            mkt_dates, mkt_prices = parse_column_pair(
                ws, col + 2, col + 3, DATA_START_ROW, ws.max_row
            )

            # Process
            result, status = process_deal(
                acq_dates, acq_prices, mkt_dates, mkt_prices,
                ann_date, deal_id, deal["deal_key"], bbg_ticker
            )

            if status == "ok":
                ts_rows.extend(result)
                success += 1
                sheet_success += 1
            elif status == "few_est":
                skip_few_est += 1
            elif status == "no_day0":
                skip_no_day0 += 1
            else:
                skip_no_data += 1

            col += COLS_PER_DEAL

        print(f" {sheet_success} deals ok")

    print(f"\n  New deals parsed: {success}")
    print(f"  Skipped (no/insufficient data): {skip_no_data}")
    print(f"  Skipped (no Day 0): {skip_no_day0}")
    print(f"  Skipped (< {MIN_EST_OBS} est obs): {skip_few_est}")

    if not ts_rows:
        print("\nNo new data to merge.")
        return

    # Merge
    new_ts = pd.concat(ts_rows, ignore_index=True)
    col_order = [
        "deal_id", "deal_key", "security_role", "security",
        "trading_date", "px_last", "ret_1d",
        "ann_date", "rel_day", "window_flag",
    ]
    new_ts = new_ts[col_order]

    combined = pd.concat([existing_ts, new_ts], ignore_index=True)
    combined.to_csv(TS_FILE, index=False)

    total_deals = combined["deal_id"].nunique()
    print(f"\n  -> Merged into {TS_FILE}")
    print(f"  Total rows: {len(combined)}")
    print(f"  Total unique deals: {total_deals}")


if __name__ == "__main__":
    main()
