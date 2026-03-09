"""
generate_bbg_excel.py  --  Generate Excel workbook with Bloomberg BDH formulas
================================================================================
Creates an Excel file where each missing deal has properly-spaced BDH formulas
that pull acquirer and S&P 500 daily prices from Bloomberg when opened in Excel
with the Bloomberg Add-in active.

BDH formulas use the "Dir","V" parameter to force VERTICAL output (dates going
down in rows), solving the column-overflow problem.

Layout per deal (6 columns):
  Col 0: deal_id header
  Col 1: [empty - BDH date output for acquirer]
  Col 2: [empty - BDH price output for acquirer]
  Col 3: [empty - BDH date output for benchmark]
  Col 4: [empty - BDH price output for benchmark]
  Col 5: gap column

After Bloomberg populates the sheet, run merge_bbg_data.py to parse the
results and merge into timeseries_long.csv.
"""

import pandas as pd
from datetime import timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import math
import sys

# ── CONFIG ──────────────────────────────────────────────────────────────────
DEALS_FILE      = "deals_master.csv"
TS_FILE         = "timeseries_long.csv"
OUTPUT_EXCEL    = "bbg_pull_missing.xlsx"
BENCHMARK_BBG   = "SPX Index"

CALENDAR_PAD_BEFORE = 380   # days before announce_date for pull start
CALENDAR_PAD_AFTER  =  20   # days after announce_date for pull end

DEALS_PER_SHEET = 40        # deals per Excel sheet (keep manageable)
COLS_PER_DEAL   = 6         # columns allocated per deal block
DATA_START_ROW  = 5         # row where BDH formula goes (1-indexed in openpyxl)

# Styles
HEADER_FONT   = Font(bold=True, size=10)
TICKER_FONT   = Font(bold=True, size=11, color="003366")
GREY_FILL     = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
BLUE_FILL     = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")


def make_bdh_formula(ticker, field, start_date, end_date):
    """Create a Bloomberg BDH formula string for Excel.

    Uses 'Dir','V' to force vertical output (dates in rows, NOT columns).
    Uses 'CDR','5D' for business days only (Mon-Fri calendar).
    """
    sd = start_date.strftime("%Y/%m/%d")
    ed = end_date.strftime("%Y/%m/%d")
    return (
        f'=BDH("{ticker}","{field}","{sd}","{ed}",'
        f'"Dir","V","CDR","5D","Fill","NA")'
    )


def main():
    print("=" * 70)
    print("  Generate Bloomberg Excel BDH Workbook")
    print("=" * 70)

    # Load deals and find missing ones
    deals = pd.read_csv(DEALS_FILE)
    deals["announce_date"] = pd.to_datetime(deals["announce_date"])

    existing_ts = pd.read_csv(TS_FILE)
    existing_deal_ids = set(existing_ts["deal_id"].unique())

    missing = deals[~deals["deal_id"].isin(existing_deal_ids)].copy()
    print(f"\nTotal deals: {len(deals)}")
    print(f"Already have data for: {len(existing_deal_ids)}")
    print(f"Missing deals to pull: {len(missing)}")

    if len(missing) == 0:
        print("No missing deals. Nothing to do.")
        return

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    n_sheets = math.ceil(len(missing) / DEALS_PER_SHEET)
    print(f"Creating {n_sheets} sheets ({DEALS_PER_SHEET} deals each)...")

    for sheet_idx in range(n_sheets):
        sheet_deals = missing.iloc[
            sheet_idx * DEALS_PER_SHEET : (sheet_idx + 1) * DEALS_PER_SHEET
        ]
        sheet_name = f"Batch_{sheet_idx + 1}"
        ws = wb.create_sheet(title=sheet_name)

        for deal_offset, (_, deal) in enumerate(sheet_deals.iterrows()):
            col_start = deal_offset * COLS_PER_DEAL + 1  # 1-indexed

            deal_id     = deal["deal_id"]
            bbg_ticker  = str(deal["acq_ticker_bbg"]).strip()
            ann_date    = deal["announce_date"]
            deal_key    = deal["deal_key"]

            # Ensure ticker ends with " Equity" for BDH
            if not bbg_ticker.endswith("Equity"):
                bbg_ticker_full = bbg_ticker + " Equity"
            else:
                bbg_ticker_full = bbg_ticker

            # Calculate date boundaries
            pull_start = ann_date - timedelta(days=CALENDAR_PAD_BEFORE)
            pull_end   = ann_date + timedelta(days=CALENDAR_PAD_AFTER)

            # ── Row 1: Deal ID ──
            cell = ws.cell(row=1, column=col_start, value=f"Deal {deal_id}")
            cell.font = HEADER_FONT
            cell.fill = BLUE_FILL

            # ── Row 2: Ticker ──
            cell = ws.cell(row=2, column=col_start, value=bbg_ticker_full)
            cell.font = TICKER_FONT

            # ── Row 3: Announce date ──
            ws.cell(row=3, column=col_start, value=f"Ann: {ann_date.strftime('%Y-%m-%d')}")

            # ── Row 4: Column headers ──
            ws.cell(row=4, column=col_start,     value="Acq Date").font = HEADER_FONT
            ws.cell(row=4, column=col_start + 1, value="Acq PX_LAST").font = HEADER_FONT
            ws.cell(row=4, column=col_start + 2, value="Mkt Date").font = HEADER_FONT
            ws.cell(row=4, column=col_start + 3, value="Mkt PX_LAST").font = HEADER_FONT
            for c in range(col_start, col_start + 4):
                ws.cell(row=4, column=c).fill = GREY_FILL

            # ── Row 5: BDH formulas ──
            # Acquirer BDH (will output dates in col_start, prices in col_start+1)
            acq_formula = make_bdh_formula(
                bbg_ticker_full, "PX_LAST", pull_start, pull_end
            )
            ws.cell(row=DATA_START_ROW, column=col_start, value=acq_formula)

            # Benchmark BDH (will output dates in col_start+2, prices in col_start+3)
            mkt_formula = make_bdh_formula(
                BENCHMARK_BBG, "PX_LAST", pull_start, pull_end
            )
            ws.cell(row=DATA_START_ROW, column=col_start + 2, value=mkt_formula)

        # Set column widths
        for deal_offset in range(len(sheet_deals)):
            col_start = deal_offset * COLS_PER_DEAL + 1
            ws.column_dimensions[_col_letter(col_start)].width = 12
            ws.column_dimensions[_col_letter(col_start + 1)].width = 14
            ws.column_dimensions[_col_letter(col_start + 2)].width = 12
            ws.column_dimensions[_col_letter(col_start + 3)].width = 14
            ws.column_dimensions[_col_letter(col_start + 4)].width = 3  # gap

    # Save
    wb.save(OUTPUT_EXCEL)
    print(f"\n  -> Saved {OUTPUT_EXCEL}")
    print(f"     {n_sheets} sheets, {len(missing)} deals total")
    print(f"\nINSTRUCTIONS:")
    print(f"  1. Open {OUTPUT_EXCEL} in Excel with Bloomberg Add-in active")
    print(f"  2. Wait for all BDH formulas to populate (may take several minutes)")
    print(f"  3. Save the file (keep same name or save-as)")
    print(f"  4. Run: .venv\\Scripts\\python.exe merge_bbg_data.py")


def _col_letter(col_num):
    """Convert 1-indexed column number to Excel letter (1=A, 27=AA, etc.)."""
    result = ""
    while col_num > 0:
        col_num, remainder = divmod(col_num - 1, 26)
        result = chr(65 + remainder) + result
    return result


if __name__ == "__main__":
    main()
