"""
generate_splc_excel.py  --  Generate Bloomberg SPLC FULL DATA Excel workbook
================================================================================
Creates an Excel file with =BDS() formulas pulling FULL supply chain data
(Suppliers + Customers) including revenue exposure, cost %, relationship info.

Each BDS FULL_DATA call returns 12 columns of data that spill to the right:
  1. Entity Ticker        2. Cost Type         3. Revenue %
  4. Cost %               5. Rel. Period       6. Rel. Year
  7. Source               8. Rel. Date         9. Rel. Count
  10. Rel. Amount         11. Company ID       12. Company Name

Layout per deal (COLS_PER_DEAL = 30 columns):
  Col 0-12:  Suppliers FULL DATA (12-col BDS spill)
  Col 13:    gap
  Col 14-26: Customers FULL DATA (12-col BDS spill)
  Col 27-29: gap

After Bloomberg populates, run merge_splc_data.py to parse into clean CSV.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import math
import os

# ── CONFIG ──────────────────────────────────────────────────────────────────
DEALS_FILE       = "data/interim/deals_master.csv"
OUTPUT_PREFIX    = "data/raw/bbg_splc_full_pull"   # will become _1.xlsx, _2.xlsx, etc.

DEALS_PER_SHEET  = 20        # fewer deals/sheet since each deal is 30 cols wide
COLS_PER_DEAL    = 30        # columns allocated per deal block
DATA_START_ROW   = 5         # row where BDS formula goes (1-indexed)
MAX_SHEETS_PER_FILE = 50     # cap per workbook to avoid Bloomberg overload

# Bloomberg BDS field configuration
SPLC_FIELDS = {
    "Suppliers": ("SUPPLY_CHAIN_SUPPLIERS_ALL_DATA", 0),    # offset within deal block
    "Customers": ("SUPPLY_CHAIN_CUSTOMERS_ALL_DATA", 14),   # offset within deal block
}

# Column headers for each 12-column BDS spill
FULL_DATA_HEADERS = [
    "Entity Ticker", "Cost Type", "Revenue %", "Cost %",
    "Rel. Period", "Rel. Year", "Source", "Rel. Date",
    "Rel. Count", "Rel. Amount", "Company ID", "Company Name"
]

# Styles
HEADER_FONT  = Font(bold=True, size=10, color="FFFFFF")
TICKER_FONT  = Font(bold=True, size=11, color="003366")
SUBHDR_FONT  = Font(bold=True, size=9)
GREY_FILL    = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
BLUE_FILL    = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
LIGHT_BLUE   = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")


def col_letter(col_num):
    """Convert 1-indexed column number to Excel letter (1=A, 27=AA, etc.)."""
    result = ""
    while col_num > 0:
        col_num, remainder = divmod(col_num - 1, 26)
        result = chr(65 + remainder) + result
    return result


def build_sheet(ws, sheet_deals):
    """Populate a single worksheet with BDS formulas for the given deals."""
    for deal_offset, (_, deal) in enumerate(sheet_deals.iterrows()):
        col_start = deal_offset * COLS_PER_DEAL + 1

        did      = deal["deal_id"]
        ticker   = str(deal["acq_ticker_bbg"]).strip()
        ann_date = deal["announce_date"]

        if not ticker.endswith("Equity"):
            ticker_full = ticker + " Equity"
        else:
            ticker_full = ticker

        # Row 1: Deal ID
        cell = ws.cell(row=1, column=col_start, value=f"Deal {did}")
        cell.font = HEADER_FONT
        cell.fill = BLUE_FILL
        for c in range(col_start + 1, col_start + COLS_PER_DEAL):
            ws.cell(row=1, column=c).fill = BLUE_FILL

        # Row 2: Ticker
        ws.cell(row=2, column=col_start, value=ticker_full).font = TICKER_FONT

        # Row 3: Announce date
        ws.cell(row=3, column=col_start,
                value=f"Ann: {ann_date.strftime('%Y-%m-%d')}")

        # Row 3-4: Field headers
        for label, (bds_field, col_offset) in SPLC_FIELDS.items():
            base_col = col_start + col_offset
            ws.cell(row=3, column=base_col, value=label).font = Font(
                bold=True, size=10, color="003366")
            for i, hdr in enumerate(FULL_DATA_HEADERS):
                cell = ws.cell(row=4, column=base_col + i, value=hdr)
                cell.font = SUBHDR_FONT
                cell.fill = LIGHT_BLUE

        # Row 5: BDS formulas
        for label, (bds_field, col_offset) in SPLC_FIELDS.items():
            formula_col = col_start + col_offset
            bds_formula = f'=BDS("{ticker_full}","{bds_field}")'
            ws.cell(row=DATA_START_ROW, column=formula_col, value=bds_formula)

    # Column widths
    for deal_offset in range(len(sheet_deals)):
        col_start = deal_offset * COLS_PER_DEAL + 1
        for c in range(col_start, col_start + COLS_PER_DEAL):
            ws.column_dimensions[col_letter(c)].width = 14


def main():
    print("=" * 70)
    print("  Generate Bloomberg BDS SPLC FULL DATA Workbooks")
    print("=" * 70)

    deals = pd.read_csv(DEALS_FILE)
    deals["announce_date"] = pd.to_datetime(deals["announce_date"])
    total_deals = len(deals)

    n_total_sheets = math.ceil(total_deals / DEALS_PER_SHEET)
    n_files = math.ceil(n_total_sheets / MAX_SHEETS_PER_FILE)

    print(f"\nTotal deals: {total_deals:,}")
    print(f"Layout: {COLS_PER_DEAL} cols/deal, {DEALS_PER_SHEET} deals/sheet")
    print(f"Max {MAX_SHEETS_PER_FILE} sheets/file -> {n_files} output files")

    deal_idx = 0
    sheet_global = 0
    saved_files = []

    for file_num in range(1, n_files + 1):
        wb = Workbook()
        wb.remove(wb.active)

        file_path = f"{OUTPUT_PREFIX}_{file_num}.xlsx"
        sheets_in_file = 0

        while sheets_in_file < MAX_SHEETS_PER_FILE and deal_idx < total_deals:
            sheet_deals = deals.iloc[deal_idx : deal_idx + DEALS_PER_SHEET]
            deal_idx += len(sheet_deals)
            sheet_global += 1
            sheets_in_file += 1

            ws = wb.create_sheet(title=f"SPLC_{sheet_global}")
            build_sheet(ws, sheet_deals)

        # Save this file
        if os.path.exists(file_path):
            try:
                os.remove(file_path)
            except Exception as e:
                print(f"  Warning: could not remove old {file_path}: {e}")

        wb.save(file_path)
        saved_files.append((file_path, sheets_in_file))
        print(f"  Saved {file_path} ({sheets_in_file} sheets)")

    print(f"\n{'='*70}")
    print(f"SUCCESS: Generated {n_files} Excel files:")
    for fp, ns in saved_files:
        print(f"  -> {fp} ({ns} sheets)")
    print(f"\n  Total: {sheet_global} sheets, {total_deals:,} deals")
    print(f"\nINSTRUCTIONS:")
    print(f"  1. Open each file ONE AT A TIME in Excel with Bloomberg Add-In")
    print(f"  2. Wait for BDS formulas to populate, then Save")
    print(f"  3. Repeat for all {n_files} files")
    print(f"{'='*70}")

if __name__ == "__main__":
    main()
