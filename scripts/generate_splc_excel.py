import pandas as pd
import math
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import os

print("="*60)
print("  GENERATING BLOOMBERG SPLC EXCEL PULLER")
print("="*60)

INPUT_DATA_PATH = "data/interim/deals_master.csv"
OUTPUT_EXCEL_PATH = "data/raw/bbg_splc_pull.xlsx"

# Number of columns reserved for ONE DEAL horizontally
COLS_PER_DEAL = 30
DEALS_PER_SHEET = 40

# Specific SPLC fields
SPLC_FIELDS = {
    "Suppliers": ("SPLC_SUPPLIERS", 1),       # Start col offset 1
    "Customers": ("SPLC_CUSTOMERS", 10),      # Start col offset 10 
    "Peers": ("SPLC_PEERS", 19)               # Start col offset 19 
}

print(f"Loading {INPUT_DATA_PATH}...")
deals_df = pd.read_csv(INPUT_DATA_PATH)

# Ensure announcement date is correctly parsed
deals_df["announce_date"] = pd.to_datetime(deals_df["announce_date"])
total_deals = len(deals_df)

print(f"Loaded {total_deals:,} deals.")

wb = Workbook()
# Remove default sheet
wb.remove(wb.active)

# Build sheets
num_sheets = math.ceil(total_deals / DEALS_PER_SHEET)
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="4F81BD")

deal_idx = 0

for sheet_num in range(num_sheets):
    ws = wb.create_sheet(title=f"SPLC_{sheet_num + 1}")
    
    # Process up to DEALS_PER_SHEET for this worksheet
    sheet_deals = deals_df.iloc[deal_idx : deal_idx + DEALS_PER_SHEET]
    deal_idx += len(sheet_deals)
    
    print(f"  Building Sheet {sheet_num + 1}/{num_sheets} for {len(sheet_deals)} deals...")
    
    col_cursor = 1
    
    for _, deal in sheet_deals.iterrows():
        did = deal["deal_id"]
        ticker = deal["acq_ticker_bbg"]
        # Format date as YYYYMMDD for Bloomberg override
        ann_date_yyyymmdd = deal["announce_date"].strftime("%Y%m%d")
        
        # Write Headers (Row 1 and 2)
        # Deal ID block
        ws.cell(row=1, column=col_cursor, value=f"DEAL HEADER: {did}").font = header_font
        ws.cell(row=1, column=col_cursor).fill = header_fill
        ws.cell(row=2, column=col_cursor, value=f"Deal ID: {did}")
        ws.cell(row=3, column=col_cursor, value=f"Ticker: {ticker}")
        ws.cell(row=4, column=col_cursor, value=f"Ann Date: {deal['announce_date'].strftime('%Y-%m-%d')}")
        
        # For each SPLC category, write exactly where it belongs in this deal's block
        for label, (bds_field, col_offset) in SPLC_FIELDS.items():
            start_col = col_cursor + col_offset - 1
            
            # Header
            ws.cell(row=1, column=start_col, value=label).font = header_font
            ws.cell(row=1, column=start_col).fill = header_fill
            
            # Sub-headers commonly returned by BDS
            ws.cell(row=2, column=start_col, value="Company Name")
            ws.cell(row=2, column=start_col+1, value="Ticker")
            ws.cell(row=2, column=start_col+2, value="Role Status")
            
            # The BDS Formula
            # Syntax: =BDS("TICKER US Equity", "SUPPLY_CHAIN_SUPPLIERS", "RELATIONSHIP_AS_OF_DATE=YYYYMMDD")
            # Bloomberg often throws "Invalid Security" if the date override is passed as two arguments instead of one.
            bds_formula = f'=BDS("{ticker}", "{bds_field}", "RELATIONSHIP_AS_OF_DATE={ann_date_yyyymmdd}")'
            ws.cell(row=3, column=start_col, value=bds_formula)
            
        col_cursor += COLS_PER_DEAL

print("\nSaving excel file...")
# Make sure previous is closed
if os.path.exists(OUTPUT_EXCEL_PATH):
    try:
        os.remove(OUTPUT_EXCEL_PATH)
    except Exception as e:
        print(f"Error removing {OUTPUT_EXCEL_PATH}: {e}")
        
wb.save(OUTPUT_EXCEL_PATH)
print("="*60)
print(f"SUCCESS: Saved {OUTPUT_EXCEL_PATH}")
print(f"It contains {num_sheets} sheets, safely 30-columns wide per deal.")
print("Open this file in a machine with Bloomberg Excel Add-In to populate SPLC data.")
